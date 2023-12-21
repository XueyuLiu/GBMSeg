import os
import shutil
import cv2
import numpy as np
import openpyxl
from tqdm import tqdm


def dice(labelpath,prepath,name):
    print(name)
    wb = openpyxl.Workbook()
    ws=wb['Sheet']
    ws.cell(row=1,column=2).value='Dice'
    ws.cell(row=1,column=1).value='name'
    num=2
    sum=[]
    dice_all=[]
    dice_1_all=[]
    #dice_2_all=[]
    #dice_3_all=[]
    #dice_4_all=[]
    n_1=0
    #n_2=0
    #n_3=0
    #n_4=0

    for root,dirs,files in os.walk(labelpath):
        for file in tqdm(files): 
            #print(file)

            s1=cv2.imread(os.path.join(labelpath,file),0)
            s2=cv2.imread(os.path.join(prepath,file),0)

            #ret,s1=cv2.threshold(pred, 0, 255, cv2.THRESH_BINARY)
            #ret,s2=cv2.threshold(target, 0, 255, cv2.THRESH_BINARY)
            try:
                row, col = s2.shape[0], s2.shape[1]
            except:
                continue
            n=0


            dice_1=0
            #dice_2=0
            #dice_3=0
            #dice_4=0

            s_1 = 0
            c_1=0
            #s_2 = 0
            #c_2=0
            #s_3 = 0
            #c_3=0
            #s_4 = 0
            #c_4=0
        
            for r in range(row):
                for c in range(col):
                    #print(s2[r][c])

                    if s1[r][c]== 255 and s2[r][c]== 255:  # 计算图像像素交集
                        s_1+=1
                    if s1[r][c]==255 or s2[r][c]==255:  # 计算图像像素并集
                        c_1+=1

                    #if s1[r][c]== 2 and s2[r][c]== 2: 
                     #   s_2+=1
                    #if s1[r][c]== 3 and s2[r][c]== 3: 
                    #    s_3+=1
                    #if s1[r][c]== 4 and s2[r][c]== 4: 
                    #    s_4+=1


                    #if s1[r][c]==2 or s2[r][c]==2: 
                     #   c_2+=1
                    #if s1[r][c]== 3 or s2[r][c]== 3: 
                     #   c_3+=1
                    #if s1[r][c]== 4 or s2[r][c]== 4: 
                    #    c_4+=1

     
            #print(s_1,c_1)
     
            if c_1==0:
                dice_1=0
            else:
                #n+=1
                dice_1=2*s_1/(c_1+s_1)
                dice_1_all.append(dice_1)

            #print(dice_1)
            #if n!=0:
                #dice_all.append((dice_1+dice_2+dice_3+dice_4)/(n))
            dice_1_all.append(dice_1)

            
            

            #d.append(2*m1/m2)
            #msg = "这是{}的dice系数".format(file) + str(dice_all)
            #print(dice_1,dice_2,dice_3,dice_4)
            #sum.append(dice_all)
            ws.cell(row=num,column=2).value=dice_1
            ws.cell(row=num,column=1).value=file
            num=num+1
            #print(msg)
    print("平均dice系数为" +str(np.mean(dice_1_all)))
    #print("类别2平均dice系数为" +str(np.mean(dice_2_all)))
    #print("类别3平均dice系数为" +str(np.mean(dice_3_all)))
    #print("类别4平均dice系数为" +str(np.mean(dice_4_all)))

    #print("平均dice系数为" +str(np.mean(dice_all)))
    #print(("平均mIoU为" +str(np.mean(dice_all)/(2-np.mean(dice_all)))))
    wb.save('/home/shiguangze/code/GBM-Seg/results/exp_results/'+name+'.xlsx')
    return str(np.mean(dice_1_all)), str(np.mean(dice_1_all)/(2-np.mean(dice_1_all)))





prepath='/home/shiguangze/code/GBM-Seg/results/exp4/'
labelpath='/home/shiguangze/code/GBM-Seg/data/test_dataset/masks'
name='0_0_1_1'
#print(prepath)
#print(labelpath)
dice_h,miou_h=dice(labelpath,os.path.join(prepath,'mask_vith_'+name),'result_4_vith_'+name)
dice_b,miou_b=dice(labelpath,os.path.join(prepath,'mask_vitb_'+name),'result_4_vitb_'+name)
dice_l,miou_l=dice(labelpath,os.path.join(prepath,'mask_vitl_'+name),'result_4_vitl_'+name)
print('dice_h='+dice_h)
print('dice_b='+dice_b)
print('dice_l='+dice_l)

print('miou_h='+miou_h)
print('miou_b='+miou_b)
print('miou_l='+miou_l)

